from PyQt6.QtWidgets import (
    QDialog, QHBoxLayout, QVBoxLayout, QLabel, QWidget, QMessageBox, 
    QTableWidget, QTableWidgetItem, QHeaderView, QAbstractItemView
)
from PyQt6.QtCore import Qt, pyqtSignal, QRectF
from PyQt6.QtGui import QPainter, QColor, QBrush, QPen
from PyQt6 import uic
from pathlib import Path
import math
from core.sql_manager import (
    get_daily_worker_progress, 
    get_daily_worker_payment_progress,
    fetch_daily_status_counts,
    fetch_today_completed_worker_stats,
    fetch_today_ev_completed_worker_stats,
    fetch_today_impossible_list,
    fetch_today_future_apply_stats,
    fetch_today_email_count,
    fetch_today_processing_list,
    fetch_today_completed_list,
    fetch_today_deferred_list,
    fetch_today_future_apply_list
)
from PyQt6.QtWidgets import QFileDialog
import pandas as pd
import os
from datetime import datetime
import pytz
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

class ClickableCard(QWidget):
    """클릭 가능한 카드 위젯"""
    clicked = pyqtSignal()

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setCursor(Qt.CursorShape.PointingHandCursor)

    def mousePressEvent(self, event):
        self.clicked.emit()
        super().mousePressEvent(event)

class PieChartWidget(QWidget):
    """파이 차트 위젯"""
    def __init__(self, data: dict, parent=None):
        super().__init__(parent)
        self.data = data  # {'label': value, ...}
        self.colors = [
            QColor("#3498db"), QColor("#e74c3c"), QColor("#2ecc71"), 
            QColor("#f1c40f"), QColor("#9b59b6"), QColor("#e67e22"), 
            QColor("#1abc9c"), QColor("#34495e")
        ]
        self.setMinimumSize(300, 200)

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing)

        rect = self.rect()
        side = min(rect.width(), rect.height())
        pie_rect = QRectF(10, 10, side - 20, side - 20)
        
        # 데이터 총합 계산
        total = sum(self.data.values())
        if total == 0:
            painter.setPen(QColor("white"))
            painter.drawText(rect, Qt.AlignmentFlag.AlignCenter, "데이터 없음")
            return

        # 파이 차트 그리기
        start_angle = 90 * 16
        color_index = 0
        
        # 범례 시작 위치 (오른쪽 여백)
        legend_x = side + 20
        legend_y = 20
        legend_available = rect.width() > side + 50

        for label, value in self.data.items():
            if value == 0: continue

            span_angle = int(-(value / total) * 360 * 16)
            
            color = self.colors[color_index % len(self.colors)]
            painter.setBrush(QBrush(color))
            painter.setPen(Qt.PenStyle.NoPen)
            
            painter.drawPie(pie_rect, start_angle, span_angle)
            
            # 범례 그리기 (공간이 있을 때만)
            if legend_available and legend_y + 20 < rect.height():
                painter.setBrush(QBrush(color))
                painter.drawRect(int(legend_x), int(legend_y), 15, 15)
                
                painter.setPen(QColor("white"))
                percentage = (value / total) * 100
                text = f"{label}: {value}건 ({percentage:.1f}%)"
                painter.drawText(int(legend_x + 25), int(legend_y + 12), text)
                
                legend_y += 25

            start_angle += span_angle
            color_index += 1

class WorkerProgressDialog(QDialog):
    """전체 업무 현황판 다이얼로그"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        
        # UI 파일 로드
        ui_path = Path(__file__).parent.parent / "ui" / "worker_progress.ui"
        uic.loadUi(str(ui_path), self)
        
        # 다이얼로그 설정
        self.setWindowTitle("전체 업무 현황판")
        self.setModal(True)
        
        # 초기화
        self._setup_ui()
        # 데이터 로드는 창이 보일 때(showEvent) 혹은 명시적으로 호출
    
    def showEvent(self, event):
        """다이얼로그가 표시될 때 데이터를 로드한다."""
        super().showEvent(event)
        self._load_overall_status()

    def _setup_ui(self):
        """UI 컴포넌트를 설정한다."""
        # 전역 스타일시트 (툴팁 스타일 포함)
        self.setStyleSheet(self.styleSheet() + """
            QToolTip {
                background-color: #2c3e50;
                color: #ecf0f1;
                border: 1px solid #3498db;
                border-radius: 4px;
                padding: 8px;
                font-size: 12px;
            }
        """)

        # 버튼 연결
        self.close_button.clicked.connect(self.accept)
        self.refresh_button.clicked.connect(self._load_overall_status)
        self.period_combo.currentTextChanged.connect(self._on_period_changed)
        self.export_report_button.clicked.connect(self._on_export_report)
        
        # 요약 정보 위젯 참조 저장소
        self.summary_widgets = {}
        self.current_chart_type = None  # 현재 표시 중인 차트 타입 (toggle 기능용)
        
        # 요약 정보 레이아웃 초기화 (Placeholders)
        self._init_summary_ui()
    
    def _on_period_changed(self, text):
        """기간 변경 시 UI 처리"""
        # 보고서 추출 버튼 가시성 제어
        self.export_report_button.setVisible(text in ["금일", "1분기"])
        
        if text == "1분기":
            # 요약 섹션 비우기
            self._clear_layout(self.summary_layout)
            self.summary_widgets.clear()
            
            # 차트 섹션 비우기
            if self.chart_container.layout():
                self._clear_layout(self.chart_container.layout())
            
            # 타이틀 업데이트
            self.title_label.setText("1분기 업무 현황")
        else:
            # 기존 UI 구조 복구 (없는 경우에만)
            if not self.summary_widgets:
                self._init_summary_ui()
            
            # 데이터 로드 (현재는 금일 데이터만 로드됨)
            self._load_overall_status()

    def _on_export_report(self):
        """보고서 추출 버튼 클릭 시 처리"""
        try:
            period = self.period_combo.currentText()
            if period != "금일":
                QMessageBox.warning(self, "보고서 추출", "현재는 '금일' 보고서만 추출 가능합니다.")
                return

            # 1. 파일 저장 경로 선택
            kst = pytz.timezone('Asia/Seoul')
            today_str = datetime.now(kst).strftime('%Y%m%d')
            default_name = f"업무현황보고서_{today_str}.xlsx"
            
            file_path, _ = QFileDialog.getSaveFileName(
                self, "보고서 저장", default_name, "Excel Files (*.xlsx)"
            )
            
            if not file_path:
                return

            # 2. 데이터 수집
            counts = fetch_daily_status_counts()
            processing_list = fetch_today_processing_list()
            completed_list = fetch_today_completed_list()
            deferred_list = fetch_today_deferred_list()
            impossible_list = fetch_today_impossible_list()
            future_apply_list = fetch_today_future_apply_list()

            # 3. Excel 파일 생성 (Pandas + openpyxl)
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                # 요약 시트
                summary_data = [
                    ["항목", "건수", "비고"],
                    ["전체 접수 (RN 고유)", counts.get('pipeline', 0), "금일 수신된 고유 RN 수"],
                    ["이메일 수신 총계", counts.get('email_pipeline', 0), "중복 포함 전체 메일 수"],
                    ["처리중", counts.get('processing', 0), "현재 작업 대기/진행 중"],
                    ["처리완료 (RNS)", counts.get('completed', 0), "작업자 상태 완료 기준"],
                    ["신청완료 (EV)", counts.get('ev_completed', 0), "EV Portal 신청 완료 기준"],
                    ["미비/보류", counts.get('deferred', 0), "서류미비, 보완요청 등"],
                    ["신청불가", counts.get('impossible', 0), "부적합, 중복 등"],
                    ["추후 신청", counts.get('future_apply', 0), "고객 요청 등으로 보류"]
                ]
                df_summary = pd.DataFrame(summary_data[1:], columns=summary_data[0])
                df_summary.to_excel(writer, sheet_name='요약', index=False)

                # 상세 시트들
                pd.DataFrame(processing_list).to_excel(writer, sheet_name='처리중', index=False)
                pd.DataFrame(completed_list).to_excel(writer, sheet_name='완료목록', index=False)
                pd.DataFrame(deferred_list).to_excel(writer, sheet_name='미비_보류목록', index=False)
                pd.DataFrame(impossible_list).to_excel(writer, sheet_name='신청불가목록', index=False)
                pd.DataFrame(future_apply_list).to_excel(writer, sheet_name='추후신청목록', index=False)

                # 스타일 적용을 위해 workbook 접근
                workbook = writer.book
                
                # 모든 시트에 대해 스타일 적용
                header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
                header_font = Font(color='FFFFFF', bold=True)
                border = Border(
                    left=Side(style='thin'), right=Side(style='thin'), 
                    top=Side(style='thin'), bottom=Side(style='thin')
                )

                for sheet_name in writer.sheets:
                    worksheet = writer.sheets[sheet_name]
                    
                    # 헤더 스타일 및 열 너비 조정
                    for col_num, col_cells in enumerate(worksheet.columns, 1):
                        # 헤더 스타일 (1행)
                        cell = worksheet.cell(row=1, column=col_num)
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        # 전체 테두리 및 자동 너비 조절
                        max_length = 0
                        for cell in col_cells:
                            cell.border = border
                            try:
                                if cell.value:
                                    length = len(str(cell.value).encode('utf-8'))
                                    if length > max_length:
                                        max_length = length
                            except:
                                pass
                        
                        adjusted_width = (max_length + 2) * 1.2
                        worksheet.column_dimensions[get_column_letter(col_num)].width = min(adjusted_width, 50)

            # 4. 완료 알림
            reply = QMessageBox.question(
                self, "보고서 추출 완료", 
                f"보고서가 성공적으로 추출되었습니다.\n경로: {file_path}\n\n파일을 지금 여시겠습니까?",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            
            if reply == QMessageBox.StandardButton.Yes:
                os.startfile(file_path)

        except Exception as e:
            QMessageBox.critical(self, "보고서 추출 실패", f"에러가 발생했습니다: {str(e)}")
            import traceback
            traceback.print_exc()

    def _init_summary_ui(self):
        """요약 UI 구조를 초기화한다."""
        self._clear_layout(self.summary_layout)
        self.summary_widgets.clear()
        
        # 초기 카드 생성 및 참조 저장
        self._create_summary_card(
            "pipeline", "파이프라인", "#3498db", 
            clickable=True, onClick=self._show_pipeline_stats,
            tooltip="고유 RN 수(중복 메일 포함 수)"
        )
        self._create_summary_card(
            "processing", "처리중", "#f1c40f",
            clickable=True, onClick=self._show_processing_list
        )
        self._create_summary_card(
            "completed", "완료", "#2ecc71", 
            clickable=True, onClick=self._show_completed_stats,
            tooltip="작업자 상태 처리완료 수(금일 EV 상 신청완료 건)"
        )
        self._create_summary_card("deferred", "미비/보류", "#e74c3c")
        self._create_summary_card("impossible", "신청불가", "#95a5a6", clickable=True, onClick=self._show_impossible_list)
        self._create_summary_card("future_apply", "추후 신청", "#9b59b6", clickable=True, onClick=self._show_future_apply_stats)
        
    def _create_summary_card(self, key: str, label: str, color: str, clickable=False, onClick=None, tooltip=None):
        """요약 카드를 생성하고 레이아웃에 추가한다."""
        if clickable:
            card = ClickableCard()
            if onClick:
                card.clicked.connect(onClick)
        else:
            card = QWidget()
        
        # 툴팁 설정
        if tooltip:
            card.setToolTip(tooltip)
        else:
            card.setToolTip(f"퇴근 시 있으면 안됌")
            
        card.setStyleSheet(f"""
            QWidget {{
                background-color: #2c3e50;
                border-radius: 10px;
                border: 1px solid #3e5871;
            }}
            QWidget:hover {{
                border: {'2px solid ' + color if clickable else '1px solid #3e5871'};
                background-color: {'#34495e' if clickable else '#2c3e50'};
            }}
        """)
        card.setMinimumHeight(110) # 아이콘 추가로 인한 높이 확보
        
        card_layout = QVBoxLayout(card)
        card_layout.setContentsMargins(12, 8, 8, 12) # 우상단 아이콘을 위해 마진 최적화
        card_layout.setSpacing(0)
        
        # [추가] 우상단 '?' 도움말 아이콘
        header_layout = QHBoxLayout()
        header_layout.addStretch()
        info_mark = QLabel("?")
        info_mark.setFixedSize(18, 18)
        info_mark.setAlignment(Qt.AlignmentFlag.AlignCenter)
        info_mark.setStyleSheet(f"""
            QLabel {{
                color: {color};
                background-color: #1a1a1a;
                border: 1px solid {color};
                border-radius: 9px;
                font-size: 11px;
                font-weight: bold;
            }}
        """)
        header_layout.addWidget(info_mark)
        card_layout.addLayout(header_layout)
        
        val_label = QLabel("0")
        val_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        val_label.setStyleSheet(f"font-size: 28px; font-weight: bold; color: {color}; border: none; background: transparent;")
        
        txt_label = QLabel(label)
        txt_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        txt_label.setStyleSheet("font-size: 14px; color: #bdc3c7; border: none; background: transparent;")
        
        card_layout.addWidget(val_label)
        card_layout.addWidget(txt_label)
        
        self.summary_layout.addWidget(card)
        
        # 값 라벨에 대한 참조 저장
        self.summary_widgets[key] = val_label
        # 카드 위젯 참조
        val_label.setProperty("card_widget", card)

    def _refresh_summary_ui(self, pipeline="0", processing="0", completed="0", deferred="0", impossible="0", future_apply="0", ev_completed="0", email_pipeline="0"):
        """요약 영역의 값을 갱신한다."""
        if "pipeline" in self.summary_widgets:
            # {pipeline}({email_pipeline}) 형식
            text = f"{pipeline}({email_pipeline})"
            self.summary_widgets["pipeline"].setText(text)
        if "processing" in self.summary_widgets:
            self.summary_widgets["processing"].setText(str(processing))
        if "completed" in self.summary_widgets:
            # {completed}({ev_completed}) 형식
            text = f"{completed}({ev_completed})"
            self.summary_widgets["completed"].setText(text)
        if "deferred" in self.summary_widgets:
            self.summary_widgets["deferred"].setText(str(deferred))
        if "impossible" in self.summary_widgets:
            self.summary_widgets["impossible"].setText(str(impossible))
        
        if "future_apply" in self.summary_widgets:
            label = self.summary_widgets["future_apply"]
            label.setText(str(future_apply))
            
            # 0이면 숨김, 1 이상이면 표시
            card = label.property("card_widget")
            if card:
                try:
                    val = int(str(future_apply).replace(',', ''))
                    card.setVisible(val > 0)
                except:
                    card.setVisible(False)
    
    def _clear_layout(self, layout):
        """레이아웃 내의 모든 위젯을 제거한다."""
        if layout:
            while layout.count():
                child = layout.takeAt(0)
                if child.widget():
                    child.widget().deleteLater()
                elif child.layout():
                    self._clear_layout(child.layout())

    def _load_overall_status(self):
        """전체 업무 현황 데이터를 로드한다."""
        try:
            # 통합 쿼리 함수 호출
            counts = fetch_daily_status_counts()
            
            self._refresh_summary_ui(
                pipeline=counts.get('pipeline', 0),
                processing=counts.get('processing', 0),
                completed=counts.get('completed', 0),
                deferred=counts.get('deferred', 0),
                impossible=counts.get('impossible', 0),
                future_apply=counts.get('future_apply', 0),
                ev_completed=counts.get('ev_completed', 0),
                email_pipeline=counts.get('email_pipeline', 0)
            )
            
            # 하단 타이틀 업데이트
            self.title_label.setText(f"실시간 업무 현황 (금일 접수: {counts.get('pipeline', 0)}건)")
            
            # 차트 컨테이너 초기화 (기본 메시지)
            self._show_message_in_chart("상단의 카드를 클릭하면 상세 통계를 볼 수 있습니다.")
            self.current_chart_type = None
            
        except Exception as e:
            print(f"Error loading status: {e}")
            self.title_label.setText("데이터 로드 실패")

    def _show_pipeline_stats(self):
        """파이프라인 건에 대한 상세 정보(이메일 수신 건수)를 보여준다 (토글 방식)."""
        if self.current_chart_type == 'pipeline':
            self._show_message_in_chart("상단의 카드를 클릭하면 상세 통계를 볼 수 있습니다.")
            self.current_chart_type = None
            return
        
        try:
            count = fetch_today_email_count()
            self._show_message_in_chart(f"금일 수신된 총 이메일 건수: {count}건")
            self.current_chart_type = 'pipeline'
        except Exception as e:
            self._show_message_in_chart(f"데이터 로드 실패: {str(e)}")
            self.current_chart_type = None

    def _show_processing_list(self):
        """처리중 건에 대한 상세 목록을 보여준다 (토글 방식)."""
        if self.current_chart_type == 'processing':
            self._show_message_in_chart("상단의 카드를 클릭하면 상세 통계를 볼 수 있습니다.")
            self.current_chart_type = None
            return
            
        try:
            items = fetch_today_processing_list()
            
            self._clear_layout(self.chart_container.layout())
            if self.chart_container.layout() is None:
                layout = QVBoxLayout()
                self.chart_container.setLayout(layout)
            
            if not items:
                self._show_message_in_chart("처리 중인 데이터가 없습니다.")
                self.current_chart_type = 'processing'
                return
            
            table = QTableWidget()
            table.setColumnCount(2)
            table.setHorizontalHeaderLabels(["RN", "상태"])
            table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
            table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
            
            self._apply_table_style(table)
            
            table.setRowCount(len(items))
            for i, item in enumerate(items):
                rn_item = QTableWidgetItem(str(item.get('RN', '')))
                rn_item.setFlags(rn_item.flags() ^ Qt.ItemFlag.ItemIsEditable)
                
                status_item = QTableWidgetItem(str(item.get('status', '')))
                status_item.setFlags(status_item.flags() ^ Qt.ItemFlag.ItemIsEditable)
                status_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                
                table.setItem(i, 0, rn_item)
                table.setItem(i, 1, status_item)
            
            self.chart_container.layout().addWidget(table)
            self.current_chart_type = 'processing'
            
        except Exception as e:
            self._show_message_in_chart(f"데이터 로드 실패: {str(e)}")
            self.current_chart_type = None

    def _show_completed_stats(self):
        """완료 건에 대한 상세 통계(파이 차트)를 보여준다 (토글 방식)."""
        if self.current_chart_type == 'completed':
            self._show_message_in_chart("상단의 카드를 클릭하면 상세 통계를 볼 수 있습니다.")
            self.current_chart_type = None
            return

        try:
            # 1. rns 테이블 기반 통계 (기존)
            stats_rns = fetch_today_completed_worker_stats()
            # 2. ev_rns 테이블 기반 통계 (신규)
            stats_ev = fetch_today_ev_completed_worker_stats()
            
            self._clear_layout(self.chart_container.layout())
            if self.chart_container.layout() is None:
                layout = QVBoxLayout()
                self.chart_container.setLayout(layout)
            
            if not stats_rns and not stats_ev:
                self._show_message_in_chart("완료된 작업 데이터가 없습니다.")
                self.current_chart_type = 'completed'
                return

            # 두 개의 차트를 가로로 배치하기 위한 레이아웃
            charts_layout = QHBoxLayout()
            
            # 왼쪽: rns 통계
            rns_wrapper = QVBoxLayout()
            rns_label = QLabel("작업자 상태 기준 (rns)")
            rns_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            rns_label.setStyleSheet("color: #ecf0f1; font-weight: bold; font-size: 14px; margin-bottom: 5px;")
            rns_wrapper.addWidget(rns_label)
            
            if stats_rns:
                chart_rns = PieChartWidget(stats_rns)
                rns_wrapper.addWidget(chart_rns)
            else:
                empty_label = QLabel("데이터 없음")
                empty_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
                empty_label.setStyleSheet("color: #bdc3c7;")
                rns_wrapper.addWidget(empty_label)
            
            # 오른쪽: ev_rns 통계
            ev_wrapper = QVBoxLayout()
            ev_label = QLabel("EV 상 완료 기준 (ev_rns)")
            ev_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            ev_label.setStyleSheet("color: #ecf0f1; font-weight: bold; font-size: 14px; margin-bottom: 5px;")
            ev_wrapper.addWidget(ev_label)
            
            if stats_ev:
                chart_ev = PieChartWidget(stats_ev)
                ev_wrapper.addWidget(chart_ev)
            else:
                empty_label = QLabel("데이터 없음")
                empty_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
                empty_label.setStyleSheet("color: #bdc3c7;")
                ev_wrapper.addWidget(empty_label)
                
            charts_layout.addLayout(rns_wrapper)
            charts_layout.addLayout(ev_wrapper)
            
            self.chart_container.layout().addLayout(charts_layout)
            self.current_chart_type = 'completed'
            
        except Exception as e:
            import traceback
            traceback.print_exc()
            self._show_message_in_chart(f"차트 로드 실패: {str(e)}")
            self.current_chart_type = None

    def _show_impossible_list(self):
        """신청불가 건에 대한 상세 목록을 보여준다 (토글 방식)."""
        if self.current_chart_type == 'impossible':
            self._show_message_in_chart("상단의 카드를 클릭하면 상세 통계를 볼 수 있습니다.")
            self.current_chart_type = None
            return
            
        try:
            items = fetch_today_impossible_list()
            
            self._clear_layout(self.chart_container.layout())
            if self.chart_container.layout() is None:
                layout = QVBoxLayout()
                self.chart_container.setLayout(layout)
            
            if not items:
                self._show_message_in_chart("신청불가 데이터가 없습니다.")
                self.current_chart_type = 'impossible'
                return
            
            table = QTableWidget()
            table.setColumnCount(3)
            table.setHorizontalHeaderLabels(["RN", "지역", "불가 사유"])
            table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
            table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
            table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
            
            self._apply_table_style(table)
            
            table.setRowCount(len(items))
            for i, item in enumerate(items):
                rn_item = QTableWidgetItem(str(item.get('RN', '')))
                rn_item.setFlags(rn_item.flags() ^ Qt.ItemFlag.ItemIsEditable)
                
                region_item = QTableWidgetItem(str(item.get('region', '')))
                region_item.setFlags(region_item.flags() ^ Qt.ItemFlag.ItemIsEditable)
                region_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                
                reason_item = QTableWidgetItem(str(item.get('reason', '')))
                reason_item.setFlags(reason_item.flags() ^ Qt.ItemFlag.ItemIsEditable)
                
                table.setItem(i, 0, rn_item)
                table.setItem(i, 1, region_item)
                table.setItem(i, 2, reason_item)
            
            self.chart_container.layout().addWidget(table)
            self.current_chart_type = 'impossible'
            
        except Exception as e:
            self._show_message_in_chart(f"데이터 로드 실패: {str(e)}")
            self.current_chart_type = None

    def _show_future_apply_stats(self):
        """추후 신청 건에 대한 지역별 집계를 보여준다 (토글 방식)."""
        if self.current_chart_type == 'future_apply':
            self._show_message_in_chart("상단의 카드를 클릭하면 상세 통계를 볼 수 있습니다.")
            self.current_chart_type = None
            return
        
        try:
            items = fetch_today_future_apply_stats()
            
            self._clear_layout(self.chart_container.layout())
            if self.chart_container.layout() is None:
                layout = QVBoxLayout()
                self.chart_container.setLayout(layout)
            
            if not items:
                self._show_message_in_chart("추후 신청 데이터가 없습니다.")
                self.current_chart_type = 'future_apply'
                return

            table = QTableWidget()
            table.setColumnCount(2)
            table.setHorizontalHeaderLabels(["지역", "개수"])
            table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
            table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
            
            self._apply_table_style(table)
            
            table.setRowCount(len(items))
            for i, item in enumerate(items):
                region_item = QTableWidgetItem(str(item.get('region', '')))
                region_item.setFlags(region_item.flags() ^ Qt.ItemFlag.ItemIsEditable)
                region_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                
                count_item = QTableWidgetItem(str(item.get('count', 0)))
                count_item.setFlags(count_item.flags() ^ Qt.ItemFlag.ItemIsEditable)
                count_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                
                table.setItem(i, 0, region_item)
                table.setItem(i, 1, count_item)
                
            self.chart_container.layout().addWidget(table)
            self.current_chart_type = 'future_apply'
            
        except Exception as e:
            self._show_message_in_chart(f"데이터 로드 실패: {str(e)}")
            self.current_chart_type = None

    def _apply_table_style(self, table):
        """테이블 공통 스타일 적용"""
        table.setStyleSheet("""
            QTableWidget {
                background-color: #2c3e50;
                gridline-color: #3e5871;
                border: none;
            }
            QHeaderView::section {
                background-color: #34495e;
                color: #ecf0f1;
                padding: 5px;
                border: 1px solid #2c3e50;
            }
            QTableWidget::item {
                color: #ecf0f1;
                padding: 5px;
            }
            QTableWidget::item:selected {
                background-color: #3498db;
            }
        """)

    def _show_message_in_chart(self, message: str):
        """차트 영역에 메시지를 표시한다."""
        # 차트 컨테이너 비우기
        if self.chart_container.layout():
            self._clear_layout(self.chart_container.layout())
        
        # 레이아웃 생성 및 설정
        if self.chart_container.layout() is None:
            layout = QVBoxLayout()
            self.chart_container.setLayout(layout)
            
        label = QLabel(message)
        label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        label.setStyleSheet("color: #bdc3c7; font-size: 14px;")
        self.chart_container.layout().addWidget(label)
    
    def _load_worker_progress(self):
        """이전 작업자 현황 로드 로직 (필요 시 현황판 하단 차트용으로 재구성 가능)"""
        pass
    
    def _show_no_data_message(self):
        pass
    
    def _show_error_message(self, message: str):
        pass
    
    def _create_chart(self, workers: list[str], counts: list[int]):
        pass

    def _create_dual_chart(self, workers: list[str], existing_counts: list[int], new_counts: list[int]):
        pass